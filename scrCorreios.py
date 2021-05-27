import sys
import re
import time
#import mysql.connector

import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# db = mysql.connector.connect(
#   host="10.20.3.6",
#   user="victordesgm",
#   passwd="covs123",
#   database="dw_covs"
# )
# mycursor = db.cursor()

def tratamentoLogradouro(logradouro):

  logradouro = logradouro.upper()
  tuplePrefixLogradouro = logradouro.partition(' ')
  prefixLogradouro = tuplePrefixLogradouro[0]

  auxLograd = logradouro.replace(prefixLogradouro, "")
  tupleSufixLograd = auxLograd.partition(' ')
  sufixLograd = tupleSufixLograd[0]

  restoLograd = auxLograd.replace(sufixLograd, "")

  if(prefixLogradouro == "ACAMPAMENTO"):
    prefixLogradouro = "ACAMP"
  elif(prefixLogradouro == "ACESSO"):
    prefixLogradouro = "AC"
  elif(prefixLogradouro == "AEROPORTO"):
    prefixLogradouro = "AER"
  elif(prefixLogradouro == "ALAMEDA"):
    prefixLogradouro = "AL"
  elif(prefixLogradouro == "AVENIDA"):
    prefixLogradouro = "AV"
  elif(prefixLogradouro == "BECO"):
    prefixLogradouro = "BC"
  elif(prefixLogradouro == "CAMINHO"):
    prefixLogradouro = "CAM"
  elif(prefixLogradouro == "CONDOMÍNIO"):
    prefixLogradouro = "COND"
  elif(prefixLogradouro == "CONDOMINIO"):
    prefixLogradouro = "COND"
  elif(prefixLogradouro == "CONJUNTO"):
    prefixLogradouro = "CJ"
  elif(prefixLogradouro == "CÓRREGO"):
    prefixLogradouro = "COR"
  elif(prefixLogradouro == "CORREGO"):
    prefixLogradouro = "COR"
  elif(prefixLogradouro == "ESCADA"):
    prefixLogradouro = "ESC"
  elif(prefixLogradouro == "ESCADARIA"):
    prefixLogradouro = "ESC"
  elif(prefixLogradouro == "ESTACIONAMENTO"):
    prefixLogradouro = "ESTAC"
  elif(prefixLogradouro == "LAGO"):
    prefixLogradouro = "LAG"
  elif(prefixLogradouro == "LARGO"):
    prefixLogradouro = "LG"
  elif(prefixLogradouro == "PASSAGEM"):
    prefixLogradouro = "PAS"
  elif(prefixLogradouro == "PASSARELA"):
    prefixLogradouro = "PSA"
  elif(prefixLogradouro == "PRAÇA"):
    prefixLogradouro = "PC"
  elif(prefixLogradouro == "RAMAL"):
    prefixLogradouro = "RAM"
  elif(prefixLogradouro == "RAMPA"):
    prefixLogradouro = "RMP"
  elif(prefixLogradouro == "RODOVIA"):
    prefixLogradouro = "ROD"
  elif(prefixLogradouro == "RÓTULA"):
    prefixLogradouro = "ROT"
  elif(prefixLogradouro == "RUA"):
    prefixLogradouro = "R"
  elif(prefixLogradouro == "RUA DE LIGAÇÃO"):
    prefixLogradouro = "R. LIG"
  elif(prefixLogradouro == "RUA DE LIGACAO"):
    prefixLogradouro = "R. LIG"
  elif(prefixLogradouro == "RUA DE LIGACÃO"):
    prefixLogradouro = "R. LIG"
  elif(prefixLogradouro == "RUA DE LIGAÇAO"):
    prefixLogradouro = "R. LIG"
  elif(prefixLogradouro == "RUA DE PEDESTRE"):
    prefixLogradouro = "R. PED"
  elif(prefixLogradouro == "RUA PARTICULAR"):
    prefixLogradouro = "R. PART"
  elif(prefixLogradouro == "RUA PROJETADA"):
    prefixLogradouro = "R. PROJ"
  elif(prefixLogradouro == "RUELA"):
    prefixLogradouro = "R"
  elif(prefixLogradouro == "TRAVESSA"):
    prefixLogradouro = "TV"
  elif(prefixLogradouro == "TRAVESSA PARTICULAR"):
    prefixLogradouro = "TV. PART"
  elif(prefixLogradouro == "TRINCHEIRA"):
    prefixLogradouro = "TCH"
  elif(prefixLogradouro == "VALA"):
    prefixLogradouro = "VAL"
  elif(prefixLogradouro == "VALE"):
    prefixLogradouro = "VLE"
  elif(prefixLogradouro == "VEREDA"):
    prefixLogradouro = "VER"
  elif(prefixLogradouro == "VIA"):
    prefixLogradouro = "V"
  elif(prefixLogradouro == "VIA EXPRESSA"):
    prefixLogradouro = "V. EXP"
  elif(prefixLogradouro == "VIA DE PEDESTRE"):
    prefixLogradouro = "V. PED"
  elif(prefixLogradouro == "VIADUTO"):
    prefixLogradouro = "VD"
  elif(prefixLogradouro == "VIELA"):
    prefixLogradouro = "VLA"
  elif(prefixLogradouro == "VILA"):
    prefixLogradouro = "VL"

  if(sufixLograd == "ADVOGADO"):
    sufixLograd = "ADV"
  if(sufixLograd == "ADVOGADA"):
    sufixLograd = "ADVA"
  if(sufixLograd == "AGENTE"):
    sufixLograd = "AG"
  if(sufixLograd == "AGRICULTOR"):
    sufixLograd = "AGRIC"
  if(sufixLograd == "AGRIMENSOR"):
    sufixLograd = "AGRIM"
  if(sufixLograd == "AJUDANTE"):
    sufixLograd = "AJ"
  if(sufixLograd == "ALMIRANTE"):
    sufixLograd = "ALM"
  if(sufixLograd == "APÓSTOLO"):
    sufixLograd = "APÓS"
  if(sufixLograd == "ARQUITETO"):
    sufixLograd = "ARQ"
  if(sufixLograd == "ARQUITETA"):
    sufixLograd = "ARQA"
  if(sufixLograd == "ARTISTA"):
    sufixLograd = "ART"
  if(sufixLograd == "AVIADOR"):
    sufixLograd = "AV"
  if(sufixLograd == "AVIADORA"):
    sufixLograd = "AVA"
  if(sufixLograd == "BARÃO"):
    sufixLograd = "BR"
  if(sufixLograd == "BARONESA"):
    sufixLograd = "BRA"
  if(sufixLograd == "BISPO"):
    sufixLograd = "BP"
  if(sufixLograd == "BRIGADEIRO"):
    sufixLograd = "BRIG"
  if(sufixLograd == "CABO"):
    sufixLograd = "CB"
  if(sufixLograd == "CAÇADOR"):
    sufixLograd = "CAC"
  if(sufixLograd == "CACIQUE"):
    sufixLograd = "CAQ"
  if(sufixLograd == "CADETE"):
    sufixLograd = "CAD"
  if(sufixLograd == "CANTOR"):
    sufixLograd = "CAN"
  if(sufixLograd == "CAPITÃO"):
    sufixLograd = "CAP"
  if(sufixLograd == "CAPITÃO-MOR"):
    sufixLograd = "CAP. MOR"
  if(sufixLograd == "CAPITÃO-TENENTE"):
    sufixLograd = "CAP. TEN"
  if(sufixLograd == "CARDEAL"):
    sufixLograd = "CARD"
  if(sufixLograd == "CAVALHEIRO"):
    sufixLograd = "CAV"
  if(sufixLograd == "COMANDANTE"):
    sufixLograd = "CNTE"
  if(sufixLograd == "COMEDIANTE"):
    sufixLograd = "COM"
  if(sufixLograd == "COMENDADOR"):
    sufixLograd = "CONDOR"
  if(sufixLograd == "CONDE"):
    sufixLograd = "CDE"
  if(sufixLograd == "CONDESSA"):
    sufixLograd = "CDESSA"
  if(sufixLograd == "CÔNEGO"):
    sufixLograd = "CON"
  if(sufixLograd == "CONSELHEIRA"):
    sufixLograd = "CONSELA"
  if(sufixLograd == "CONSELHEIRO"):
    sufixLograd = "CONSEL"
  if(sufixLograd == "CORONEL"):
    sufixLograd = "CEL"
  if(sufixLograd == "DENTISTA"):
    sufixLograd = "DENT"
  if(sufixLograd == "DEPUTADA"):
    sufixLograd = "DEPA"
  if(sufixLograd == "DEPUTADO"):
    sufixLograd = "DEP"
  if(sufixLograd == "DESEMBARGADOR"):
    sufixLograd = "DESEMB"
  if(sufixLograd == "DOM"):
    sufixLograd = "D"
  if(sufixLograd == "DONA"):
    sufixLograd = "DN"
  if(sufixLograd == "DOUTOR"):
    sufixLograd = "DR"
  if(sufixLograd == "DOUTORA"):
    sufixLograd = "DRA"
  if(sufixLograd == "DUQUE"):
    sufixLograd = "DQ"
  if(sufixLograd == "DUQUESA"):
    sufixLograd = "DQA"
  if(sufixLograd == "EMBAIXADOR"):
    sufixLograd = "EMB"
  if(sufixLograd == "EMBAIXATRIZ"):
    sufixLograd = "EMBA"
  if(sufixLograd == "ENFERMEIRO"):
    sufixLograd = "ENF"
  if(sufixLograd == "ENGENHEIRA"):
    sufixLograd = "ENGA"
  if(sufixLograd == "ENGENHEIRO"):
    sufixLograd = "ENG"
  if(sufixLograd == "ESTUDANTE"):
    sufixLograd = "EST"
  if(sufixLograd == "FREI"):
    sufixLograd = "FR"
  if(sufixLograd == "FREIRE"):
    sufixLograd = "FRE"
  if(sufixLograd == "GENERAL"):
    sufixLograd = "GAL"
  if(sufixLograd == "GOVERNADOR"):
    sufixLograd = "GOV"
  if(sufixLograd == "GRÃO"):
    sufixLograd = "GR"
  if(sufixLograd == "IMPERADOR"):
    sufixLograd = "IMP"
  if(sufixLograd == "IMPERATRIZ"):
    sufixLograd = "IMPA"
  if(sufixLograd == "IRMÃO"):
    sufixLograd = "IR"
  if(sufixLograd == "IRMÃ"):
    sufixLograd = "IR"
  if(sufixLograd == "JORNALISTA"):
    sufixLograd = "JORN"
  if(sufixLograd == "JUNIOR"):
    sufixLograd = "JR"
  if(sufixLograd == "LORDE"):
    sufixLograd = "LD"
  if(sufixLograd == "MADRE"):
    sufixLograd = "MDE"
  if(sufixLograd == "MAESTRO"):
    sufixLograd = "MTRO"
  if(sufixLograd == "MAJOR"):
    sufixLograd = "MJ"
  if(sufixLograd == "MARECHAL"):
    sufixLograd = "MCHAL"
  if(sufixLograd == "MARQUÊS"):
    sufixLograd = "MRQ"
  if(sufixLograd == "MARQUESA"):
    sufixLograd = "MRQA"
  if(sufixLograd == "MESTRE"):
    sufixLograd = "MTRE"
  if(sufixLograd == "MINISTRO"):
    sufixLograd = "MIN"
  if(sufixLograd == "MISSÍONÁRIO"):
    sufixLograd = "MIS"
  if(sufixLograd == "MONGE"):
    sufixLograd = "MG"
  if(sufixLograd == "MONSENHOR"):
    sufixLograd = "MONSR"
  if(sufixLograd == "NOSSA SENHORA"):
    sufixLograd = "N. SRA"
  if(sufixLograd == "OUVIDOR"):
    sufixLograd = "OUV"
  if(sufixLograd == "PADRE"):
    sufixLograd = "PE"
  if(sufixLograd == "PAPA"):
    sufixLograd = "PA"
  if(sufixLograd == "PASTOR"):
    sufixLograd = "PAS"
  if(sufixLograd == "POETA"):
    sufixLograd = "PTA"
  if(sufixLograd == "PREFEITA"):
    sufixLograd = "PREFA"
  if(sufixLograd == "PREFEITO"):
    sufixLograd = "PREF"
  if(sufixLograd == "PRESIDENTE"):
    sufixLograd = "PRES"
  if(sufixLograd == "PRINCESA"):
    sufixLograd = "PSA"
  if(sufixLograd == "PRÍNCIPE"):
    sufixLograd = "PRIN"
  if(sufixLograd == "PROCURADOR"):
    sufixLograd = "PROC"
  if(sufixLograd == "PROCURADORA"):
    sufixLograd = "PROCA"
  if(sufixLograd == "PROFESSOR"):
    sufixLograd = "PROF"
  if(sufixLograd == "PROFESSOR-DOUTOR"):
    sufixLograd = "PROF. DR"
  if(sufixLograd == "PROFESSORA"):
    sufixLograd = "PROFA"
  if(sufixLograd == "PROFETA"):
    sufixLograd = "PFT"
  if(sufixLograd == "PROMOTOR"):
    sufixLograd = "PROM"
  if(sufixLograd == "RABINO"):
    sufixLograd = "RAB"
  if(sufixLograd == "RADIALISTA"):
    sufixLograd = "RAD"
  if(sufixLograd == "RAINHA"):
    sufixLograd = "RHA"
  if(sufixLograd == "REI"):
    sufixLograd = "REI"
  if(sufixLograd == "REVERENDO"):
    sufixLograd = "VER"
  if(sufixLograd == "SANTA"):
    sufixLograd = "STA"
  if(sufixLograd == "SANTO"):
    sufixLograd = "STO"
  if(sufixLograd == "SANTISSÍMA"):
    sufixLograd = "STMA"
  if(sufixLograd == "SÃO"):
    sufixLograd = "S"
  if(sufixLograd == "SEGUNDO-SARGENTO"):
    sufixLograd = "SEG. SARG"
  if(sufixLograd == "SENADOR"):
    sufixLograd = "SEN"
  if(sufixLograd == "SENADORA"):
    sufixLograd = "SENA"
  if(sufixLograd == "SENHOR"):
    sufixLograd = "SR"
  if(sufixLograd == "SENHORITA"):
    sufixLograd = "STA"
  if(sufixLograd == "SINHÁ"):
    sufixLograd = "SHA"
  if(sufixLograd == "SOLDADO"):
    sufixLograd = "SOL"
  if(sufixLograd == "SUB-TENENTE"):
    sufixLograd = "S. TEN"
  if(sufixLograd == "TENENTE"):
    sufixLograd = "TEN"
  if(sufixLograd == "TENENTE-CORONEL"):
    sufixLograd = "TEN. CEL"
  if(sufixLograd == "TENENTE-SARGENTO"):
    sufixLograd = "TEN. SARG"
  if(sufixLograd == "TERCEIRO-SARGENTO"):
    sufixLograd = "TERC. SARG"
  if(sufixLograd == "VEREADOR"):
    sufixLograd = "VER"
  if(sufixLograd == "VIGÁRIO"):
    sufixLograd = "VIG"
  if(sufixLograd == "VISCONDE"):
    sufixLograd = "VISC"
  if(sufixLograd == "VISCONDESSA"):
    sufixLograd = "VISCA"
  if(sufixLograd == "VOLUNTÁRIA"):
    sufixLograd = "VOLTA"
  if(sufixLograd == "VOLUNTÁRIO"):
    sufixLograd = "VOLT"

  logradouro = prefixLogradouro + " " + sufixLograd + " " + restoLograd
  logradouro = logradouro.strip()
  logradouro = re.sub(' +', ' ', logradouro)
  return logradouro

def correiosCaller(contEncontrados, contNaoEncontrados, contTotalBuscaCorreios, indiceBaseNova, countNumCepsSPGFC,
                   workbookNova, workbookNotFound, workbookBaseSPGFC, sheetNova, sheetNotFound, sheetBaseSPGFC):
  try:
    driver = webdriver.Chrome(executable_path='chromedriver.exe')
    driver.get("https://buscacepinter.correios.com.br/app/endereco/index.php")

    inicioBusca = contEncontrados + contNaoEncontrados
    for i in range(inicioBusca, contTotalBuscaCorreios):
      fluxoNormalAdicao = False
      cep_end_elem = WebDriverWait(driver, 120).until(
        EC.presence_of_element_located((By.XPATH, "//input[@name='endereco']"))
      )
      cep_end_elem.clear()
      cep_end_elem.send_keys(cepsSiteCorreios[i])

      btn_pesquisar_elem = WebDriverWait(driver, 120).until(
        EC.presence_of_element_located((By.XPATH, "//button[@name='btn_pesquisar']"))
      )
      btn_pesquisar_elem.click()

      # Caso o CEP exista serão extraidas essas informações
      try:
        td_logradouro_elem = WebDriverWait(driver, 10).until(
          EC.presence_of_element_located((By.XPATH, "//td[@data-th='Logradouro/Nome']"))
        )

        td_cep_elem = WebDriverWait(driver, 10).until(
          EC.presence_of_element_located((By.XPATH, "//td[@data-th='CEP']"))
        )

        td_bairro_elem = WebDriverWait(driver, 10).until(
          EC.presence_of_element_located((By.XPATH, "//td[@data-th='Bairro/Distrito']"))
        )

        td_localidade_elem = WebDriverWait(driver, 10).until(
          EC.presence_of_element_located((By.XPATH, "//td[@data-th='Localidade/UF']"))
        )
        # Subistituindo o traço por vazio nos ceps e tratando o logradouro
        ReplaceDashCep = td_cep_elem.text.replace('-', '')
        logTratado = tratamentoLogradouro(td_logradouro_elem.text)

        # Populando a planilha de encontrados com informações do cep encontrado no site dos correios
        linhaColunaCEPBaseNova = "A" + str(indiceBaseNova)
        sheetNova[linhaColunaCEPBaseNova] = re.sub(' ', '', ReplaceDashCep)

        linhaColunaLograBaseNova = "B" + str(indiceBaseNova)
        sheetNova[linhaColunaLograBaseNova] = logTratado

        linhaColunaIdBaseNova = "C" + str(indiceBaseNova)
        sheetNova[linhaColunaIdBaseNova] = None

        linhaColunaBairroBaseNova = "D" + str(indiceBaseNova)
        sheetNova[linhaColunaBairroBaseNova] = td_bairro_elem.text.upper()

        linhaColunaLocalidadeBaseNova = "E" + str(indiceBaseNova)
        sheetNova[linhaColunaLocalidadeBaseNova] = td_localidade_elem.text.upper()

        workbookNova.save(filename="Ceps Encontrados.xlsx")
        indiceBaseNova += 1

        # Populando a planilha da Fundac (SPGFC)
        linhaColunaCEBaseSPGFC = "A" + str(countNumCepsSPGFC+2)
        sheetBaseSPGFC[linhaColunaCEBaseSPGFC] = re.sub(' ', '', ReplaceDashCep)

        linhaColunaLograBaseSPGFC = "B" + str(countNumCepsSPGFC+2)
        sheetBaseSPGFC[linhaColunaLograBaseSPGFC] = logTratado

        linhaColunaIdBaseSPGFC = "C" + str(countNumCepsSPGFC+2)
        sheetBaseSPGFC[linhaColunaIdBaseSPGFC] = countNumCepsSPGFC+1

        linhaColunaBairroBaseSPGFC = "D" + str(countNumCepsSPGFC+2)
        sheetBaseSPGFC[linhaColunaBairroBaseSPGFC] = td_bairro_elem.text.upper()

        linhaColunaLocalidadeBaseSPGFC = "E" + str(countNumCepsSPGFC+2)
        sheetBaseSPGFC[linhaColunaLocalidadeBaseSPGFC] = td_localidade_elem.text.upper()

        workbookBaseSPGFC.save(filename="BASE_RESERVA_CEPS_FUNDAC.xlsx")
        countNumCepsSPGFC += 1
        contEncontrados += 1

        # Inserindo valores encontrados no banco de dados
        # try:
        #   mycursor.execute("INSERT INTO tab_ceps_spgeo (cep, logradouro) VALUES (%s, %s)", (re.sub(' ', '', ReplaceDashCep), logTratado))
        #   db.commit()
        # except:
        #   print("Não foi possível salvar CEP no banco de dados")
      except Exception as e:
        # Populando nova planilha com informações do cep não encontrado
        linhaColunaCEP = "A" + str(contNaoEncontrados+2)
        sheetNotFound[linhaColunaCEP] = cepsSiteCorreios[i]
        workbookNotFound.save(filename="Ceps Não Encontrados.xlsx")
        print("\nCEP: ", cepsSiteCorreios[i], "não encontrado")
        contNaoEncontrados += 1

      btn_voltar_elem = WebDriverWait(driver, 120).until(
        EC.presence_of_element_located((By.XPATH, "//button[@name='btn_voltar']"))
      )
      btn_voltar_elem.click()
      fluxoNormalAdicao = True
      sys.stdout.write("\rContagem de CEPS pesquisados no Correios: %i" %i)
      sys.stdout.flush()
      time.sleep(1)
  except Exception as e:
    print(e)
    print("\nOcorreu um erro ao acessar o site, a busca recomeçará em breve")
    if(fluxoNormalAdicao == False):

      # exclui a ultima linha adicionada na planilha
      linhaColunaNotFound = "A" + str(contNaoEncontrados+2)
      sheetNotFound[linhaColunaNotFound] = None
      contNaoEncontrados -= 1
    time.sleep(60)
    driver.close()
    correiosCaller(contEncontrados, contNaoEncontrados, contTotalBuscaCorreios, indiceBaseNova, countNumCepsSPGFC,
                   workbookNova, workbookNotFound, workbookBaseSPGFC, sheetNova, sheetNotFound, sheetBaseSPGFC)

  print("CEPS encontrados no site dos correios: ", contEncontrados)
  print("CEPS não encontrados no site dos correios: ", contNaoEncontrados)
  driver.close()

print("\t\t\t\t\t\t\t\tAPLICAÇÃO DE LIMPEZA DE PLANILHA DE CEPS - SPGEO")

try:
  # Abrindo planilha com CEPS sujos a serem tratados
  workbookBaseSuja = load_workbook(filename="CEPS_LOGS_LIMPEZA.xlsx")
  sheetBaseSuja = workbookBaseSuja.active
  cepsBaseSuja = []
  logsBaseSuja = []
  idsBaseSuja = []

  # Populando vetor de logradouro e vetor de cep com o que está na base suja
  i = 2
  while(True):
    cepsBaseSuja.append(sheetBaseSuja.cell(row = i, column=1).value)
    if(cepsBaseSuja[i-2] == None):
      cepsBaseSuja.pop()
      break
    cepsBaseSuja[i-2] = str(cepsBaseSuja[i-2])
    if(len(cepsBaseSuja[i-2]) == 6):
      cepsBaseSuja[i-2] = "0" + cepsBaseSuja[i-2] + "0"
    elif(len(cepsBaseSuja[i-2]) == 7):
      cepsBaseSuja[i-2] = "0" + cepsBaseSuja[i-2]

    logsBaseSuja.append(sheetBaseSuja.cell(row = i, column=2).value)
    idsBaseSuja.append(sheetBaseSuja.cell(row = i, column=3).value)

    # Tratamentos a serem realizados com os logradouros
    logsBaseSuja[i-2] = str(logsBaseSuja[i-2])
    logsBaseSuja[i-2] = tratamentoLogradouro(logsBaseSuja[i-2])

    i += 1
except Exception as e:
  print(e)
  print("Não foi possível abrir a base para ser realizada a limpeza")
finally:
  print("Base suja aberta com sucesso!!!")

# Abrindo planilha com a base CEM
try:
  workbookBaseCEM = load_workbook(filename="BASE_CEM.xlsx")
  sheetBaseCEM = workbookBaseCEM.active
  cepsBaseCEM = []
  logsBaseCEM = []
  idsBaseCEM = []

  # Populando vetor de logradouro e vetor de cep com o que está na base CEM
  i = 2
  while(True):
    cepsBaseCEM.append(sheetBaseCEM.cell(row = i, column=1).value)

    if(cepsBaseCEM[i-2] == None):
      cepsBaseCEM.pop()
      break

    cepsBaseCEM[i-2] = str(cepsBaseCEM[i-2])

    if(len(cepsBaseCEM[i-2]) == 6):
      cepsBaseCEM[i-2] = "0" + cepsBaseCEM[i-2] + "0"
    elif(len(cepsBaseCEM[i-2]) == 7):
      cepsBaseCEM[i-2] = "0" + cepsBaseCEM[i-2]

    logsBaseCEM.append(sheetBaseCEM.cell(row = i, column=2).value)
    idsBaseCEM.append(sheetBaseCEM.cell(row = i, column=3).value)
    # Tratamentos a serem realizados com os logradouros
    logsBaseCEM[i-2] = str(logsBaseCEM[i-2])
    logsBaseCEM[i-2] = tratamentoLogradouro(logsBaseCEM[i-2])

    i += 1
except Exception as e:
  print(e)
  print("Não foi possível abrir a base CEM")
finally:
  print("Base CEM aberta com sucesso!!!")

# Abrindo planilha SPGFC
try:
  workbookBaseSPGFC = load_workbook(filename="BASE_RESERVA_CEPS_FUNDAC.xlsx")
  sheetBaseSPGFC = workbookBaseSPGFC.active
  cepsBaseSPGFC = []
  logsBaseSPGFC = []
  idsBaseSPGFC = []

  # Populando vetor de logradouro e vetor de cep com o que está na base SPGFC
  i = 2
  countNumCepsSPGFC = 0
  while(True):
    cepsBaseSPGFC.append(sheetBaseSPGFC.cell(row = i, column=1).value)

    if(cepsBaseSPGFC[i-2] == None):
      cepsBaseSPGFC.pop()
      break

    cepsBaseSPGFC[i-2] = str(cepsBaseSPGFC[i-2])
    cepsBaseSPGFC[i-2].replace('-', '')
    cepsBaseSPGFC[i-2].replace(' ', '')

    if(len(cepsBaseSPGFC[i-2]) == 6):
      cepsBaseSPGFC[i-2] = "0" + cepsBaseSPGFC[i-2] + "0"
    elif(len(cepsBaseSPGFC[i-2]) == 7):
      cepsBaseSPGFC[i-2] = "0" + cepsBaseSPGFC[i-2]

    logsBaseSPGFC.append(sheetBaseSPGFC.cell(row = i, column=2).value)
    idsBaseSPGFC.append(sheetBaseSPGFC.cell(row = i, column=3).value)
    # Tratamentos a serem realizados com os logradouros
    logsBaseSPGFC[i-2] = str(logsBaseSPGFC[i-2])
    logsBaseSPGFC[i-2] = tratamentoLogradouro(logsBaseSPGFC[i-2])

    countNumCepsSPGFC += 1
    i += 1
except Exception as e:
  print(e)
  print("Não foi possível abrir a base reserva da Fundac")
finally:
  print("Base Fundac (SPGFC) aberta com sucesso!!!")

# Criando nova planilha com os CEPS encontrados
workbookNova = Workbook()
sheetNova  = workbookNova.active
sheetNova["A1"] = "CEP"
sheetNova["B1"] = "Logradouro"
sheetNova["C1"] = "Id"
sheetNova["D1"] = "Marcação de Verificação"
workbookNova.save(filename="Ceps Encontrados.xlsx")

# Criando planilha com os ceps não encontrados
workbookNotFound = Workbook()
sheetNotFound  = workbookNotFound.active
sheetNotFound["A1"] = "CEP"
sheetNotFound["B1"] = "Logradouro"
sheetNotFound["C1"] = "Id"
workbookNotFound.save(filename="Ceps Não Encontrados.xlsx")
# Criar o vetor para receber os valores que não forem encontrados
cepsNotFound = []
logsNotFound = []
idsNotFound = []

# Comparação de CEPS entre base suja e base CEM
indiceBaseNova = 2
indiceBaseNotFound = 2
contadorEncontrados = 0
contadorNaoEncontrados = 0
encontrado = False

try:
  for i in range(len(cepsBaseSuja)):
    for j in range(len(cepsBaseCEM)):
      if(cepsBaseSuja[i] == cepsBaseCEM[j]):

        linhaColunaCEPBaseNova = "A" + str(indiceBaseNova)
        sheetNova[linhaColunaCEPBaseNova] = cepsBaseCEM[j]

        linhaColunaLograBaseNova = "B" + str(indiceBaseNova)
        sheetNova[linhaColunaLograBaseNova] = logsBaseCEM[j]

        linhaColunaIdBaseNova = "C" + str(indiceBaseNova)
        sheetNova[linhaColunaIdBaseNova] = idsBaseSuja[i]

        # Logradouro não bate
        if(logsBaseSuja[i] != logsBaseCEM[j]):
          linhaColunaMarcacaoBaseNova = "D" + str(indiceBaseNova)
          sheetNova[linhaColunaMarcacaoBaseNova] = "Logradouro Difere: Base CEM"
          workbookNova.save(filename="Ceps Encontrados.xlsx")

        contadorEncontrados += 1
        indiceBaseNova += 1
        encontrado = True
        break

    sys.stdout.write("\rComparando dados da base suja com a base CEM: %i" %i)
    sys.stdout.flush()

    if(encontrado == False):
      cepsNotFound.append(cepsBaseSuja[i])
      logsNotFound.append(logsBaseSuja[i])
      idsNotFound.append(idsBaseSuja[i])

      contadorNaoEncontrados += 1

    encontrado = False
except Exception as e:
  print(e)
  print("\nNão foi possivel fazer a comparação entre a base suja e a base CEM")

print("\nCEPS encontrados na base CEM: ", contadorEncontrados)
print("CEPS não encontrados na base CEM: ", contadorNaoEncontrados)

print("Comparando dados da base suja com a base Reserva CEPS Fundac...")
#Comparação entre BaseNotFound e Base Fundac (SPGFC)
cepsSiteCorreios = []
logsSiteCorreios = []
idsSiteCorreios = []
contadorEncontrados = 0
contadorNaoEncontrados = 0

try:
  for i in range(len(cepsNotFound)):
    for j in range(len(cepsBaseSPGFC)):
      if(cepsNotFound[i] == cepsBaseSPGFC[j]):
        linhaColunaCEPBaseNova = "A" + str(indiceBaseNova)
        sheetNova[linhaColunaCEPBaseNova] = cepsBaseSPGFC[j]

        linhaColunaLograBaseNova = "B" + str(indiceBaseNova)
        sheetNova[linhaColunaLograBaseNova] = logsBaseSPGFC[j]

        linhaColunaIdBaseNova = "C" + str(indiceBaseNova)
        sheetNova[linhaColunaIdBaseNova] = None

        # Logradouro não bate
        if(logsNotFound[i] != logsBaseSPGFC[j]):
          linhaColunaMarcacaoBaseNova = "D" + str(indiceBaseNova)
          sheetNova[linhaColunaMarcacaoBaseNova] = "Log Difere: Base Reserva Fundac"
          workbookNova.save(filename="Ceps Encontrados.xlsx")

        contadorEncontrados += 1
        indiceBaseNova += 1
        encontrado = True
        break

    sys.stdout.write("\rComparando dados da base suja com a base Reserva Fundac: %i" %i)
    sys.stdout.flush()

    if(encontrado == False):
      cepsSiteCorreios.append(cepsBaseSuja[i])
      logsSiteCorreios.append(logsBaseSuja[i])
      idsSiteCorreios.append(idsBaseSuja[i])

      contadorNaoEncontrados += 1

    encontrado = False
except Exception as e:
  print(e)
  print("\nNão foi possivel fazer a comparação entre a base suja e a base Reserva CEPS Fundac")

print("\nCEPS encontrados na base Reserva CEPS Fundac: ", contadorEncontrados)
print("CEPS não encontrados na base Reserva CEPS Fundac: ", contadorNaoEncontrados)

print('Iniciando busca no site dos correios...')
contTotalBuscaCorreios = contadorNaoEncontrados
contEncontrados = 0
contNaoEncontrados = 0
correiosCaller(contEncontrados, contNaoEncontrados, contTotalBuscaCorreios, indiceBaseNova, countNumCepsSPGFC,
                   workbookNova, workbookNotFound, workbookBaseSPGFC, sheetNova, sheetNotFound, sheetBaseSPGFC)
